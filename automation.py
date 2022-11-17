from openpyxl import Workbook, load_workbook
import sys
input_file_name = sys.argv[1]
reset_values=sys.argv[3]
wb = load_workbook('/var/lib/jenkins/workspace/Automation/Testsuite.xlsx')
sh=wb['TestSuite'] 
wb2=load_workbook(input_file_name)
sh2 = wb2['Testcases']
test_folders_loc=[]
for row in sh['B']:
        test_folders_loc.append(row.value)
print(test_folders_loc)
test_suite_val1=[]
for col in sh['A']:
        test_suite_val1.append(col.value)
print(test_suite_val1)
if reset_values=="True":
    current_row=2
    for eachRow in sh.iter_rows():
        sh.cell(row=current_row,column=1).value=0
        current_row+=1
    wb.save('/var/lib/jenkins/workspace/Automation/Testsuite.xlsx')
    current_row2=2
    for eachRow2 in sh2.iter_rows():
        sh2.cell(row=current_row2,column=1).value=0
        current_row2+=1
    wb2.save(input_file_name)
else: exit
if(input_file_name in test_folders_loc):
    t1=test_folders_loc.index(input_file_name)+1
    sh['A'+str(t1)].value=1
    print(sh['A'+str(t1)].value)
    wb.save('/var/lib/jenkins/workspace/Automation/Testsuite.xlsx')
if input_file_name=='/var/lib/jenkins/workspace/Automation/person_dim_test_suite.xlsx':
   input_product_name = sys.argv[2]
   if input_product_name=='telecom-dev':
      sh2['A2'].value=sh2['A3'].value=sh2['A4'].value=1
      wb2.save(input_file_name)
   elif input_product_name=="mobi-dev":
        sh2['A5'].value=sh2['A6'].value=sh2['A7'].value=1
        wb2.save(input_file_name)
   elif input_product_name=="rivermine-dev":
        sh2['A8'].value=sh2['A9'].value=sh2['A10'].value=sh2['A11'].value=1
        wb2.save(input_file_name)
   else:
    print('input product name is invalid')
else: exit
        

